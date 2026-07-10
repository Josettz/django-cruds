from django.contrib import messages
from django.http import HttpResponse
from django.shortcuts import redirect
from django.urls import reverse
from django.utils import timezone


class GroupRequiredMixin:
    """
    Mixin que verifica si el usuario pertenece a uno o varios grupos (roles).
    Si no pertenece a ninguno de los grupos requeridos, redirige con error.
    """
    group_required = []
    group_redirect_url = '/'
    group_error_message = 'No tienes permiso para acceder a esta sección.'

    def dispatch(self, request, *args, **kwargs):
        if not request.user.is_authenticated:
            return self.handle_no_permission()
        if request.user.is_superuser:
            return super().dispatch(request, *args, **kwargs)
        if not self.group_required:
            return super().dispatch(request, *args, **kwargs)
        user_groups = set(request.user.groups.values_list('name', flat=True))
        if not any(g in user_groups for g in self.group_required):
            messages.error(request, self.group_error_message)
            return redirect(self.group_redirect_url)
        return super().dispatch(request, *args, **kwargs)

    def handle_no_permission(self):
        from django.contrib.auth.views import redirect_to_login
        return redirect_to_login(self.request.get_full_path())


class StaffRequiredMixin:
    """
    Mixin que verifica si el usuario es miembro del staff.
    Si no es staff, redirige con mensaje de error.

    Uso:
        class BrandDeleteView(LoginRequiredMixin, StaffRequiredMixin, DeleteView):
            ...

    ¿POR QUÉ?
    Porque solo el personal autorizado (staff) debe poder
    eliminar registros. Un usuario normal puede ver y crear,
    pero no borrar información importante del sistema.

    ¿CÓMO FUNCIONA?
    1. El usuario intenta acceder a una vista protegida
    2. dispatch() se ejecuta ANTES que la vista
    3. Si user.is_staff es False → redirige con mensaje de error
    4. Si user.is_staff es True → ejecuta la vista normalmente
    """

    # URL a donde redirigir si no es staff
    staff_redirect_url = '/'
    staff_error_message = 'You do not have permission to perform this action. Staff access required.'

    def dispatch(self, request, *args, **kwargs):
        """
        dispatch() es el primer método que se ejecuta en una CBV.
        Interceptamos aquí para verificar permisos ANTES de
        procesar la petición (GET o POST).
        """
        if not request.user.is_staff:
            messages.error(request, self.staff_error_message)
            return redirect(self.staff_redirect_url)

        return super().dispatch(request, *args, **kwargs)


class ExportListMixin:
    """
    Mixin genérico para exportar el listado (queryset filtrado) de una
    ListView a PDF o Excel.

    Uso:
        class ProductListView(ExportListMixin, ListView):
            export_filename = 'productos'
            export_title = 'Listado de Productos'
            export_fields = [
                ('Name', 'name'),
                ('Brand', 'brand.name'),        # ruta con puntos
                ('Suppliers', lambda o: ', '.join(s.name for s in o.suppliers.all())),  # callable
            ]

    ¿CÓMO FUNCIONA?
    1. El usuario hace clic en un botón -> GET con ?export=pdf|excel
       (se conservan TODOS los filtros porque van en la misma querystring).
    2. get() detecta el parámetro y exporta self.get_queryset() (ya filtrado).
    3. Cada columna de export_fields se resuelve: ruta con puntos, callable,
       o método sin argumentos.
    """

    # Lista de (encabezado, accessor). accessor = str | "a.b.c" | callable(obj)
    export_fields = []
    export_filename = 'export'
    export_title = None  # encabezado del PDF; por defecto usa export_filename

    def get_export_queryset(self):
        """Queryset a exportar. Por defecto, el ya filtrado de la vista."""
        return self.get_queryset()

    def get_export_fields(self):
        """
        Fuente ÚNICA de columnas a exportar -> [(encabezado, accessor)].
        Por defecto usa el atributo estático export_fields, pero las vistas
        pueden sobreescribirlo para devolver solo las columnas visibles
        (misma config que usa la tabla del listado).
        """
        return self.export_fields

    def _resolve(self, obj, accessor):
        if callable(accessor):
            return accessor(obj)
        value = obj
        for part in accessor.split('.'):
            value = getattr(value, part)
            if callable(value):
                value = value()
        return value

    def get_export_rows(self):
        fields = self.get_export_fields()
        for obj in self.get_export_queryset():
            yield [self._resolve(obj, acc) for _, acc in fields]

    def get_export_headers(self):
        return [header for header, _ in self.get_export_fields()]

    def get(self, request, *args, **kwargs):
        export = request.GET.get('export')
        if export == 'excel':
            return self.export_excel()
        if export == 'pdf':
            return self.export_pdf()
        return super().get(request, *args, **kwargs)

    # --- EXCEL ---
    def export_excel(self):
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill

        wb = Workbook()
        ws = wb.active
        ws.title = (self.export_title or self.export_filename)[:31]

        headers = self.get_export_headers()
        ws.append(headers)
        head_font = Font(bold=True, color='FFFFFF')
        head_fill = PatternFill('solid', fgColor='343A40')
        for cell in ws[1]:
            cell.font = head_font
            cell.fill = head_fill

        widths = [len(str(h)) for h in headers]
        for row in self.get_export_rows():
            clean = ['' if v is None else v for v in row]
            ws.append(clean)
            for i, v in enumerate(clean):
                widths[i] = max(widths[i], len(str(v)))

        # auto-ancho proporcional al contenido (acotado 10..60)
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = min(60, max(10, w + 2))

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{self._fname("xlsx")}"'
        wb.save(response)
        return response

    # --- PDF ---
    def export_pdf(self):
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape, portrait
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.lib.enums import TA_CENTER
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{self._fname("pdf")}"'

        headers = self.get_export_headers()
        ncols = len(headers) or 1

        # orientación dinámica: pocas columnas -> vertical, muchas -> horizontal
        pagesize = portrait(A4) if ncols <= 4 else landscape(A4)
        # tamaño de fuente dinámico
        if ncols <= 5:
            font_size = 9
        elif ncols <= 8:
            font_size = 7
        else:
            font_size = 6

        margin = 1 * cm
        doc = SimpleDocTemplate(
            response, pagesize=pagesize,
            leftMargin=margin, rightMargin=margin, topMargin=margin, bottomMargin=margin,
        )

        styles = getSampleStyleSheet()
        cell_style = ParagraphStyle('cell', parent=styles['Normal'], fontSize=font_size, leading=font_size + 2)
        head_style = ParagraphStyle('head', parent=cell_style, textColor=colors.white,
                                    fontName='Helvetica-Bold')
        elements = [
            Paragraph(self.export_title or self.export_filename, styles['Title']),
            Spacer(1, 0.4 * cm),
        ]

        # datos como Paragraph -> permiten wrap automático
        data = [[Paragraph(str(h), head_style) for h in headers]]
        for row in self.get_export_rows():
            data.append([Paragraph('' if v is None else str(v), cell_style) for v in row])

        # ancho proporcional que llena todo el espacio disponible (centrado/aprovechado)
        avail = pagesize[0] - 2 * margin
        col_widths = [avail / ncols] * ncols

        table = Table(data, colWidths=col_widths, repeatRows=1, hAlign='CENTER')
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#343A40')),
            ('FONTSIZE', (0, 0), (-1, -1), font_size),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F2F2F2')]),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('LEFTPADDING', (0, 0), (-1, -1), 4),
            ('RIGHTPADDING', (0, 0), (-1, -1), 4),
            ('TOPPADDING', (0, 0), (-1, -1), 3),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ]))
        elements.append(table)
        doc.build(elements)
        return response

    def _fname(self, ext):
        stamp = timezone.localtime().strftime('%Y%m%d_%H%M')
        return f'{self.export_filename}_{stamp}.{ext}'


class DynamicColumnsMixin:
    """
    Selección dinámica de columnas para una ListView, persistida en sesión.
    Fuente ÚNICA de columnas (tabla + PDF + Excel) = atributo COLUMNS:
        COLUMNS = [
            {'key': 'name', 'label': 'Nombre', 'default': True, 'accessor': 'name'},
            ...
        ]
    Provee también get_export_fields() (para ExportListMixin) filtrando por las
    columnas visibles, y agrega al contexto: columns, visible, visible_count,
    total_columns y querystring (filtros sin page/columns).
    """

    COLUMNS = []
    columns_session_key = None  # por defecto: "<app>.<model>_columns"

    def get_columns_session_key(self):
        if self.columns_session_key:
            return self.columns_session_key
        return f'{self.model._meta.label_lower}_columns'

    def get_default_columns(self):
        return [c['key'] for c in self.COLUMNS if c.get('default')]

    def get_visible_columns(self):
        """Keys visibles. Fuente: GET (al aplicar) -> sesión -> defaults."""
        all_keys = [c['key'] for c in self.COLUMNS]
        g = self.request.GET
        key = self.get_columns_session_key()

        if 'reset_columns' in g:
            self.request.session.pop(key, None)
            return self.get_default_columns()

        if 'columns' in g:
            selected = [k for k in g.getlist('columns') if k in all_keys]
            if not selected:  # mínimo obligatorio: 1 columna
                selected = self.get_default_columns()
            self.request.session[key] = selected
            return selected

        saved = self.request.session.get(key)
        if saved:
            valid = [k for k in saved if k in all_keys]
            return valid or self.get_default_columns()

        return self.get_default_columns()

    def get_export_fields(self):
        """Exporta SOLO columnas visibles, en el orden de COLUMNS."""
        visible = set(self.get_visible_columns())
        return [(c['label'], c['accessor']) for c in self.COLUMNS if c['key'] in visible]

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        visible = self.get_visible_columns()
        ctx['visible'] = visible
        ctx['columns'] = [
            {'key': c['key'], 'label': c['label'], 'checked': c['key'] in visible}
            for c in self.COLUMNS
        ]
        ctx['visible_count'] = len(visible)
        ctx['total_columns'] = len(self.COLUMNS)

        params = self.request.GET.copy()
        for k in ('page', 'columns', 'reset_columns'):
            params.pop(k, None)
        ctx['querystring'] = params.urlencode()
        ctx['filters'] = self.request.GET
        return ctx


class GenericFilterMixin:
    """
    Filtrado declarativo para una ListView mediante el atributo filters:
        filters = [
            {'name': 'name',  'label': 'Nombre', 'type': 'text', 'field': 'name'},
            {'name': 'email', 'label': 'Email',  'type': 'text', 'field': 'email'},
            {'name': 'is_active', 'label': 'Activo', 'type': 'boolean'},
        ]
    Tipos: 'text' (icontains), 'boolean' (0/1). Aplica en get_queryset y
    expone 'filter_configs' al contexto para renderizar el formulario genérico.
    """

    filters = []

    def apply_filters(self, qs):
        g = self.request.GET
        for f in self.filters:
            val = g.get(f['name'], '').strip()
            if not val:
                continue
            if f['type'] == 'text':
                qs = qs.filter(**{f.get('field', f['name']) + '__icontains': val})
            elif f['type'] == 'boolean' and val in ('0', '1'):
                qs = qs.filter(**{f.get('field', f['name']): val == '1'})
        return qs

    def get_queryset(self):
        return self.apply_filters(super().get_queryset())

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        g = self.request.GET
        ctx['filter_configs'] = [
            {**f, 'value': g.get(f['name'], '')} for f in self.filters
        ]
        return ctx


class ListFeaturesMixin(DynamicColumnsMixin, GenericFilterMixin, ExportListMixin):
    """
    Combo reutilizable para listados genéricos: columnas dinámicas + filtros
    declarativos + exportación PDF/Excel + filas y URLs listas para la plantilla
    genérica `billing/_generic_list.html`.

    Atributos a definir en la vista:
        model, COLUMNS, filters, paginate_by
        export_filename, export_title
        page_title, object_label
        list_url_name, create_url_name, detail_url_name,
        update_url_name, delete_url_name
    """

    page_title = None
    object_label = 'registro'
    list_url_name = None
    create_url_name = None
    detail_url_name = None
    update_url_name = None
    delete_url_name = None

    def _row_url(self, name, pk):
        return reverse(name, args=[pk]) if name else None

    def _img_cell(self, url):
        """Convierte una URL de imagen en miniatura HTML (segura) para la tabla.
        Sin imagen -> placeholder no-image.svg, igual que en productos."""
        from django.utils.safestring import mark_safe
        from django.utils.html import escape
        from django.templatetags.static import static
        src = escape(url) if url else static('billing/no-image.svg')
        return mark_safe(
            f'<img src="{src}" class="rounded" '
            f'style="width:48px;height:48px;object-fit:cover;" alt="">'
        )

    def _bool_cell(self, val):
        """Convierte un valor Sí/No (o booleano) en un badge de color para la tabla."""
        from django.utils.safestring import mark_safe
        truthy = str(val).strip().lower() in ('sí', 'si', 'true', '1', 'yes', 'activo')
        if truthy:
            return mark_safe('<span class="badge bg-success">Activo</span>')
        return mark_safe('<span class="badge bg-secondary">Inactivo</span>')

    def _cell_value(self, obj, col):
        val = self._resolve(obj, col['accessor'])
        if col.get('image'):
            return self._img_cell(val)
        if col.get('boolean'):
            return self._bool_cell(val)
        return val

    def get_display_rows(self, object_list):
        visible = self.get_visible_columns()
        cols = [c for c in self.COLUMNS if c['key'] in visible]
        rows = []
        for obj in object_list:
            rows.append({
                'pk': obj.pk,
                'cells': [self._cell_value(obj, c) for c in cols],
                'detail_url': self._row_url(self.detail_url_name, obj.pk),
                'update_url': self._row_url(self.update_url_name, obj.pk),
                'delete_url': self._row_url(self.delete_url_name, obj.pk),
            })
        return rows

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        visible = ctx['visible']
        ctx['visible_columns'] = [c for c in self.COLUMNS if c['key'] in visible]
        ctx['rows'] = self.get_display_rows(ctx['object_list'])
        ctx['page_title'] = self.page_title or self.model._meta.verbose_name_plural.title()
        ctx['object_label'] = self.object_label
        ctx['list_url'] = reverse(self.list_url_name) if self.list_url_name else ''
        ctx['create_url'] = reverse(self.create_url_name) if self.create_url_name else ''
        return ctx


class GenericDetailMixin:
    """
    DetailView genérica para `billing/_generic_detail.html`.
    Define detail_fields = [(label, accessor), ...] y los url names.
    """

    detail_fields = []
    page_title = None
    list_url_name = None
    update_url_name = None
    delete_url_name = None

    def _resolve(self, obj, accessor):
        if callable(accessor):
            return accessor(obj)
        value = obj
        for part in accessor.split('.'):
            value = getattr(value, part)
            if callable(value):
                value = value()
        return value

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        obj = self.object
        ctx['detail_rows'] = [
            {'label': label, 'value': self._resolve(obj, acc)}
            for label, acc in self.detail_fields
        ]
        # Flags para el detalle estilo "ficha" (imagen a la izquierda + badge):
        # la imagen y el estado activo se renderizan aparte, no como fila.
        field_names = {f.name for f in self.model._meta.get_fields()}
        ctx['has_image'] = 'image' in field_names
        ctx['has_active'] = 'is_active' in field_names
        ctx['page_title'] = self.page_title or self.model._meta.verbose_name.title()
        ctx['list_url'] = reverse(self.list_url_name) if self.list_url_name else ''
        ctx['update_url'] = reverse(self.update_url_name, args=[obj.pk]) if self.update_url_name else ''
        ctx['delete_url'] = reverse(self.delete_url_name, args=[obj.pk]) if self.delete_url_name else ''
        return ctx
